from openpyxl import load_workbook
class Write(object):
    def __init__(self):
        self.path_excel = r'data/test.xlsx'
        self.wb = load_workbook(self.path_excel)

    def write(self):
        # 选择工作表
        sheets = self.wb['test']
        path = u'testcs.csv'
        with open(path, 'w') as f:
            for i in range(1,420):
                left=sheets.cell(row=i, column=1).value
                right = sheets.cell(row=i, column=2).value
                f.write(str(left)+','+str(right)+'\n')
            f.flush()

if __name__=='__main__':
    ob=Write()
    ob.write()
    print('write done.')