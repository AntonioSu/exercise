#coding=utf-8
from openpyxl import load_workbook
class Write(object):
    def __init__(self):
        self.path_excel = r'data/test.xlsx'
        self.wb = load_workbook(self.path_excel)
    def write(self,result):
        # 选择工作表
        sheets = self.wb['test']
        for i,res in enumerate(result):
            sheets.cell(row=i+2, column=2).value =res
        self.wb.save(self.path_excel)




